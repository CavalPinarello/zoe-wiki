#!/usr/bin/env python3
import openpyxl
import json

wb = openpyxl.load_workbook('/Users/martinkawalski/Downloads/ZoE_Hackathon_Complete_Data.xlsx', data_only=True)
complete_data = {}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Skip first two rows (title and empty row), get headers from row 3
    headers = []
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=3, column=col)
        if cell.value:
            headers.append(str(cell.value).strip())
    
    # Get data starting from row 4
    sheet_data = []
    for row in range(4, sheet.max_row + 1):
        row_data = {}
        has_data = False
        for col in range(1, min(len(headers) + 1, sheet.max_column + 1)):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                row_data[headers[col-1]] = str(cell.value).strip()
                has_data = True
        if has_data:
            sheet_data.append(row_data)
    
    complete_data[sheet_name] = {
        'headers': headers,
        'data': sheet_data
    }
    print(f'{sheet_name}: {len(headers)} columns, {len(sheet_data)} rows')

# Save to JSON
with open('complete_excel_data.json', 'w') as f:
    json.dump(complete_data, f, indent=2)

print('\nData saved to complete_excel_data.json')

# Print sample data for verification
for sheet_name, sheet_info in complete_data.items():
    if sheet_info['data']:
        print(f'\n{sheet_name} - First row:')
        print(json.dumps(sheet_info['data'][0], indent=2)[:500])